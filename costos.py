import win32com.client
import pandas as pd
from datetime import datetime
import subprocess
import re
import os
import sys
import time
from openpyxl import load_workbook, Workbook
import pyperclip


# Conexiones para poder entrar a SAP GUI
try:
    sapGuiAuto = win32com.client.GetObject('SAPGUI')
    application = sapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)
except Exception as e:
    print(f"Recuerda primero abrir SAP en la transacción Z2L, vuelve a correr el comando")
    time.sleep(3)
    input("Presiona Enter para continuar.")
    sys.exit(1)


now = datetime.now()
dt_string = now.strftime("%d_%m_%Y_%H_%M")

# Obtener la ruta del escritorio
desktop = os.path.join(os.path.expanduser('~'), 'OneDrive - BASF' , 'Desktop')
username = os.getlogin()

# Definir el nombre de la carpeta
nombreArchivo = "Pop Savings 2023 W32.xlsx" # Cambiar a voluntad

# Nombre del material
nombre = ""

# Unir la ruta del escritorio con el nombre de la carpeta
folderdir = os.path.join(desktop, nombreArchivo)

# Fechas
lastYear = now.year - 1
startDate=f'{now.day}.{now.month}.{lastYear}'
endDate=f'{now.day}.{now.month}.{now.year}'

def main():
    print('''

 /$$$$$$$  /$$                                                   /$$       /$$          
| $$__  $$|__/                                                  |__/      | $$          
| $$  \ $$ /$$  /$$$$$$  /$$$$$$$  /$$    /$$ /$$$$$$  /$$$$$$$  /$$  /$$$$$$$  /$$$$$$ 
| $$$$$$$ | $$ /$$__  $$| $$__  $$|  $$  /$$//$$__  $$| $$__  $$| $$ /$$__  $$ /$$__  $$
| $$__  $$| $$| $$$$$$$$| $$  \ $$ \  $$/$$/| $$$$$$$$| $$  \ $$| $$| $$  | $$| $$  \ $$
| $$  \ $$| $$| $$_____/| $$  | $$  \  $$$/ | $$_____/| $$  | $$| $$| $$  | $$| $$  | $$
| $$$$$$$/| $$|  $$$$$$$| $$  | $$   \  $/  |  $$$$$$$| $$  | $$| $$|  $$$$$$$|  $$$$$$/
|_______/ |__/ \_______/|__/  |__/    \_/    \_______/|__/  |__/|__/ \_______/ \______/ 
                                                                                        
                                                                                        
                                                                                        

''')
    print(f"Hola {username}, excelente día \n")
    print("========================================================================")
    time.sleep(1)
    print("Se procederá a leer la tabla de Excel para poder sacar los codigos SAP e ir llenando los precios de cada mes \n")
    print("========================================================================")


    costos(folderdir)


def costos(folderdir):
    excelSheet = "BOTAutomation"

    try:
        # Guardar en un dataframe
        df = pd.read_excel(folderdir, header=0, sheet_name=excelSheet)

    except Exception as e:
        print(f"Error {e} es muy probable de que el excel esté abierto, cierralo para que se pueda analizar")
        input("Presiona enter para continua")
    
    # Se obtiene la lista de batches
    batches = df["CODE"]

    try:
        for index, batch in enumerate(batches):

            session.findById("wnd[0]")
            session.findById("wnd[0]/tbar[0]/okcd").text = "CKM3N"
            session.findById("wnd[0]").sendVKey(0)
            session.findById("wnd[0]/usr/ctxtMLKEY-MATNR").text = batch
            session.findById("wnd[0]/usr/ctxtMLKEY-WERKS_ML_PRODUCTIVE").text = "MX31"
            session.findById("wnd[0]/usr/ctxtMLKEY-WERKS_ML_PRODUCTIVE").setFocus()
            session.findById("wnd[0]/usr/ctxtMLKEY-WERKS_ML_PRODUCTIVE").caretPosition = 4
            session.findById("wnd[0]").sendVKey(0)
            session.findById("wnd[0]/usr/btn%#AUTOTEXT003").press()
            session.findById("wnd[0]/usr/cmbMLKEY-CURTP").setFocus()
            session.findById("wnd[0]/usr/cmbMLKEY-CURTP").key = "40"
            
            # Ir iterando por meses
            for i in range(1, 13):
                session.findById("wnd[0]/usr/txtMLKEY-POPER").text = i
                session.findById("wnd[0]/usr/txtMLKEY-BDATJ").text = "2023"
                session.findById("wnd[0]").sendVKey(0)

                price = session.findById("/app/con[0]/ses[0]/wnd[0]/usr/txtCKMLCR-PVPRS").text

                df.iloc[index, i] = price

            # Regresar para poner otro o terminar
            session.findById("wnd[0]/tbar[0]/btn[15]").press()

    except Exception as e:
        print(f"Ocurrió un error {e}")

    # Cargar el archivo Excel
    book = load_workbook(folderdir)
    sheet = book[excelSheet]
    
    # Empezamos a escribir desde la segunda fila para sobreescribir datos, manteniendo el encabezado
    start_row = 2

    # Escribir el df en el archivo a partir de la segunda fila
    for index, row in df.iterrows():
        for col_num, value in enumerate(row.values, 1):
            sheet.cell(row=start_row + index, column=col_num, value=value)
    
    # Guardar los cambios y cerrar el libro
    book.save(folderdir)

    print("Listo, consulta el excel.")
    time.sleep(3)

if __name__ == "__main__":
    main()